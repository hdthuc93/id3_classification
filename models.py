import openpyxl


class Id3Tree:
    def __init__(self):
        self.name = ""
        self.type_table = {}


class LoggingExcel:
    def __init__(self):
        self.path = ""
        self.col = 1
        self.row = 2
        self.book = openpyxl.Workbook()
        self.sheet = self.book.active
        self.sheet.cell(row=1, column=1).value = "start"

    def append(self, val):
        self.sheet.cell(row=self.row, column=self.col).value = str(val)

    def writeToFile(self, path):
        self.book.save(path)
        self.book.close()