import openpyxl
import codecs


def read_excel(path):
    workbook = openpyxl.load_workbook(path)
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)
    lst_input = []

    for rowVal in worksheet.iter_rows():
        lst_input.append([cellVal.value for cellVal in rowVal])

    workbook.close()
    return lst_input


def write_result_id3(tree, path):

    lst_string = []
    prepare_string(tree, "", lst_string)

    out = codecs.open(path, "w", "utf-8")
    out.write("\n".join(lst_string))

    out.close()
    return


def prepare_string(tree, string, lst_string):
    string += tree.name

    for key, val in tree.type_table.items():
        local_string = string + "()" + key + "()"
        if type(val) == str:
            lst_string.append(local_string + val)
        else:
            prepare_string(val, local_string, lst_string)


def read_txt_file(path):
    lst =[]
    with codecs.open(path, "r", "utf-8") as fin:
        for line in fin:
            lst.append(line.replace("\n", ""))

    # print(lst)
    return lst


def write_excel_file(lst, path):
    book = openpyxl.Workbook()
    sheet = book.active
    row = 1
    for child_lst in lst:
        col = 1
        for item in child_lst:
            sheet.cell(row=row, column=col).value = str(item)
            col += 1
        row += 1

    book.save(path)
    book.close()
    return